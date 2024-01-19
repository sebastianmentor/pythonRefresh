import time
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

LOGG_PATH = 'my_logg.txt'
EXCEL_PATH = 'login_history.xlsx'


def log_to_excel(nr_of_loggins:int):
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        dt = datetime.now()
        ws.append([nr_of_loggins, dt.year, dt.month, dt.day ,dt.hour, dt.second])
        wb.save(EXCEL_PATH)

    except FileNotFoundError as e:
        print(e + '\n Allt gick åt helvete')
        


def try_read_logg(path:str, logg_list:list):
    for nr_of_try in range(1, 4):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                for line in f:
                    logg_list.append(line)

            log_to_excel(len(logg_list))
            break

        except Exception as e:
            print(e)
            print(f"Fail on try {nr_of_try}")

        time.sleep(0.5) 


if not os.path.exists(EXCEL_PATH):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Antal loggin'
    ws['B1'] = 'År'
    ws['C1'] = 'Månad'
    ws['D1'] = 'Dag'
    ws['E1'] = 'Timme'
    ws['F1'] = 'Sekund'

    wb.save(EXCEL_PATH)




if __name__ == "__main__":
    while True:
        time.sleep(30)
        new_loggs = []
        try_read_logg(LOGG_PATH,new_loggs)

        if new_loggs:

            with open(LOGG_PATH, 'w') as f:
                pass

            with open('My_database.txt', 'a', encoding='utf-8') as f:
                for log in new_loggs:
                    f.write(log)